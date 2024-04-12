import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
import openpyxl

custom_headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36',
    'Accept-Language': 'vi-VN,vi;q=0.9',
    'Accept-Encoding': 'gzip, deflate, br, zstd',
    'Accept': '*/*',
    'Referer': 'https://www.amazon.com/'
}

visited_urls = set()
number = 0
max = 0

def get_product_info(url):
    response = requests.get(url, headers=custom_headers)
    if response.status_code != 200:
        print(f"Error in getting webpage: {url}")
        return None

    soup = BeautifulSoup(response.text, "lxml")

    # title
    title_element = soup.select_one("#productTitle")
    title = title_element.text.strip() if title_element else None

    # price
    price_element = soup.select_one('span.a-offscreen')
    price = price_element.text if price_element else None

    # rating
    rating_element = soup.select_one("#acrPopover")
    rating_text = rating_element.attrs.get("title") if rating_element else None
    rating = rating_text.replace("out of 5 stars", "") if rating_text else None

    # image_url
    image_element = soup.select_one("#landingImage")
    image = image_element.attrs.get("src") if image_element else None

    # description
    description_element = soup.select_one("#productDescription")
    description = description_element.text.strip() if description_element else None

    return {
        "title": title,
        "price": price,
        "rating": rating,
        "image": image,
        "description": description,
        "url": url
    }

def parse_listing(listing_url):
    global visited_urls
    global number
    global max

    response = requests.get(listing_url, headers=custom_headers)
    soup_search = BeautifulSoup(response.text, "lxml")
    link_elements = soup_search.select("[data-asin] h2 a")
    page_data = []

    for link in link_elements:
        full_url = urljoin(listing_url, link.attrs.get("href"))
        if full_url not in visited_urls:
            if max != 'all' and number == int(max):
                break

            visited_urls.add(full_url)
            print(f"#{number+1} Scraping product from {full_url[:100]}", flush=True)
            product_info = get_product_info(full_url)
            print(product_info)
            if product_info:
                page_data.append(product_info)
            number += 1

    if max == 'all' or number < int(max):
        next_page_el = soup_search.select_one('a.s-pagination-next')
        if next_page_el:
            next_page_url = next_page_el.attrs.get('href')
            next_page_url = urljoin(listing_url, next_page_url)
            page_data += parse_listing(next_page_url)

    return page_data

def main():
    global max
    search_key = input("What type of product do you wanna search?: ")
    max_product = input("How many products do you wanna search? (Enter 'all' if u wanna find all products): ")
    max = max_product

    search_url = f"https://www.amazon.com/s?k={search_key}"
    data = parse_listing(search_url)

    # Tạo một workbook mới
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Ghi tiêu đề (header) vào hàng đầu tiên
    worksheet.append(['Title', 'Price', 'Rating', 'Image URL', 'Description', 'Product URL'])

    # Ghi dữ liệu vào các hàng tiếp theo
    for product in data:
        worksheet.append([
            product['title'],
            product['price'],
            product['rating'],
            product['image'],
            product['description'],
            product['url']
        ])

    # Sắp xếp dữ liệu theo giá tăng dần
    sorted_data = sorted(data, key=lambda x: (x['price'] or float('inf')))

    # Tạo một worksheet mới để chứa dữ liệu đã sắp xếp
    sorted_worksheet = workbook.create_sheet('Sorted by Price')
    sorted_worksheet.append(['Title', 'Price', 'Rating', 'Image URL', 'Description', 'Product URL'])

    # Ghi dữ liệu đã sắp xếp vào worksheet mới
    for product in sorted_data:
        sorted_worksheet.append([
            product['title'],
            product['price'],
            product['rating'],
            product['image'],
            product['description'],
            product['url']
        ])


    # Tạo một worksheet mới để chứa dữ liệu đã sắp xếp theo rating
    sorted_by_rating_worksheet = workbook.create_sheet('Sorted by Rating')
    sorted_by_rating_worksheet.append(['Title', 'Price', 'Rating', 'Image URL', 'Description', 'Product URL'])

    # Sắp xếp dữ liệu theo rating giảm dần
    sorted_by_rating_data = sorted(data, key=lambda x: (float(x['rating'] or 0) if x['rating'] else 0), reverse=True)

    # Ghi dữ liệu đã sắp xếp theo rating vào worksheet mới
    for product in sorted_by_rating_data:
        sorted_by_rating_worksheet.append([
            product['title'],
            product['price'],
            product['rating'],
            product['image'],
            product['description'],
            product['url']
        ])

    workbook.save(f'/content/drive/My Drive/GG_Colab/Crawl_Amazon/{search_key}_data.xlsx')

if __name__ == '__main__':
    main()
